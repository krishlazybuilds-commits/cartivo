import csv
import io
from decimal import Decimal

from django.db.models import Avg, Count
from django.http import StreamingHttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.filters import OrderingFilter
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAdminUser, IsAuthenticated, IsAuthenticatedOrReadOnly
from rest_framework.response import Response

from .filters import PostgresSearchFilter
from .models import Category, Product, ProductImage, ProductVariant, Review, Warehouse, WarehouseStock, WishlistItem
from .validators import validate_import_file
from .serializers import (
    CategorySerializer,
    ProductImageSerializer,
    ProductImportSerializer,
    ProductSerializer,
    ProductVariantSerializer,
    ReviewSerializer,
    WarehouseSerializer,
    WarehouseStockSerializer,
    WishlistItemSerializer,
)


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    lookup_field = "slug"
    search_fields = ("name",)
    ordering_fields = ("name", "created_at")

    def get_permissions(self):
        if self.action in ("list", "retrieve"):
            return [IsAuthenticatedOrReadOnly()]
        return [IsAdminUser()]


class ProductViewSet(viewsets.ModelViewSet):
    serializer_class = ProductSerializer
    lookup_field = "slug"
    filter_backends = (DjangoFilterBackend, PostgresSearchFilter, OrderingFilter)
    filterset_fields = ("category", "is_active", "is_featured", "is_new", "on_sale")
    search_fields = ("name", "description", "sku")
    ordering_fields = ("price", "created_at", "name")

    def get_queryset(self):
        qs = Product.objects.select_related("category").prefetch_related("variants", "images").annotate(
            avg_rating=Avg("reviews__rating"),
            review_count=Count("reviews"),
        )
        # Only staff (who manage the catalog) may see inactive/draft products;
        # the public catalog is limited to active ones. Prevents leaking
        # unreleased products (and their stock/sku) via the public API.
        user = self.request.user
        if not (user and user.is_staff):
            qs = qs.filter(is_active=True)

        # Allow filtering by a comma-separated list of product IDs (e.g.
        # ?ids=1,2,3). Used by the frontend checkout to validate guest cart
        # prices against real server-side values.
        ids_param = self.request.query_params.get("ids", "").strip()
        if ids_param:
            id_list = [int(x) for x in ids_param.split(",") if x.strip().isdigit()]
            if id_list:
                qs = qs.filter(id__in=id_list)

        # Explicit, deterministic ordering (newest first, id as tiebreaker) so
        # pagination is stable. The annotations add a GROUP BY, which otherwise
        # makes the model's default ordering non-guaranteed for the paginator.
        return qs.order_by("-created_at", "id")

    def get_permissions(self):
        if self.action in ("list", "retrieve", "related"):
            return [IsAuthenticatedOrReadOnly()]
        return [IsAdminUser()]

    @action(detail=True, methods=["get"])
    def related(self, request, slug=None):
        """Return up to 4 related products from the same category,
        excluding the current product, ordered by highest-rated first."""
        product = self.get_object()
        related_qs = (
            Product.objects
            .filter(category=product.category)
            .exclude(pk=product.pk)
            .select_related("category")
            .prefetch_related("variants", "images")
            .annotate(
                avg_rating=Avg("reviews__rating"),
                review_count=Count("reviews"),
            )
            .order_by("-avg_rating", "-created_at")[:4]
        )
        serializer = self.get_serializer(related_qs, many=True)
        return Response(serializer.data)

    EXPORT_FIELDS = [
        "id", "name", "slug", "sku", "category_name",
        "price", "sale_price", "stock", "description",
        "is_active", "is_featured", "is_new", "on_sale", "badge",
    ]

    def _product_rows(self):
        qs = self.get_queryset().select_related("category")
        for p in qs:
            yield {f: getattr(p, f, "") for f in self.EXPORT_FIELDS}

    @action(detail=False, methods=["get"])
    def export(self, request):
        fmt = request.query_params.get("format", "csv")

        if fmt == "xlsx":
            return self._export_xlsx()
        return self._export_csv()

    def _export_csv(self):
        rows = list(self._product_rows())

        def stream():
            buffer = io.StringIO()
            writer = csv.DictWriter(buffer, fieldnames=self.EXPORT_FIELDS)
            writer.writeheader()
            yield buffer.getvalue()
            for row in rows:
                buffer = io.StringIO()
                row_out = {k: str(v) if v is not None else "" for k, v in row.items()}
                writer = csv.DictWriter(buffer, fieldnames=self.EXPORT_FIELDS)
                writer.writerow(row_out)
                yield buffer.getvalue()

        response = StreamingHttpResponse(stream(), content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="products.csv"'
        return response

    def _export_xlsx(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(self.EXPORT_FIELDS)
        for row in self._product_rows():
            ws.append([str(row[f]) if row[f] is not None else "" for f in self.EXPORT_FIELDS])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from django.http import HttpResponse
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="products.xlsx"'
        return response

    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser, FormParser])
    def import_products(self, request):
        ser = ProductImportSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        file = ser.validated_data["file"]

        # Validate the file is genuinely CSV or XLSX (magic bytes, content-type,
        # encoding).  This is defense-in-depth — processing a non-CSV/XLSX file
        # would produce empty results, not code execution, but early rejection
        # is cleaner and logs suspicious activity.
        validate_import_file(file)

        ext = file.name.rsplit(".", 1)[-1].lower() if "." in file.name else ""
        if ext == "xlsx":
            return self._import_xlsx(file)
        return self._import_csv(file)

    def _import_csv(self, file):
        decoded = file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))
        return self._process_rows(reader)

    def _import_xlsx(self, file):
        from openpyxl import load_workbook

        wb = load_workbook(file, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = [str(c).strip().lower() if c else "" for c in next(rows_iter, [])]
        rows = [dict(zip(header, [str(c) if c is not None else "" for c in row])) for row in rows_iter]
        return self._process_rows(rows)

    def _process_rows(self, rows):
        created = 0
        updated = 0
        errors = []

        for i, row in enumerate(rows, start=2):
            sku = (row.get("sku") or "").strip()
            name = (row.get("name") or "").strip()
            if not sku or not name:
                errors.append(f"Row {i}: missing sku or name")
                continue

            price_raw = (row.get("price") or "").strip()
            try:
                price = Decimal(price_raw) if price_raw else Decimal("0")
            except Exception:
                errors.append(f"Row {i}: invalid price '{price_raw}'")
                continue

            cat_name = (row.get("category_name") or "").strip()
            category = None
            if cat_name:
                category = Category.objects.filter(name__iexact=cat_name).first()

            defaults = {
                "name": name,
                "price": price,
                "description": (row.get("description") or "").strip(),
                "is_active": (row.get("is_active") or "").strip().lower() in ("1", "true", "yes"),
                "is_featured": (row.get("is_featured") or "").strip().lower() in ("1", "true", "yes"),
                "is_new": (row.get("is_new") or "").strip().lower() in ("1", "true", "yes"),
                "on_sale": (row.get("on_sale") or "").strip().lower() in ("1", "true", "yes"),
                "badge": (row.get("badge") or "").strip(),
            }

            sale_price_raw = (row.get("sale_price") or "").strip()
            if sale_price_raw:
                try:
                    defaults["sale_price"] = Decimal(sale_price_raw)
                except Exception:
                    errors.append(f"Row {i}: invalid sale_price '{sale_price_raw}'")
                    continue

            stock_raw = (row.get("stock") or "").strip()
            try:
                defaults["stock"] = int(stock_raw) if stock_raw else 0
            except ValueError:
                errors.append(f"Row {i}: invalid stock '{stock_raw}'")
                continue

            if category:
                defaults["category"] = category

            _, is_created = Product.objects.update_or_create(sku=sku, defaults=defaults)
            if is_created:
                created += 1
            else:
                updated += 1

        return Response({
            "created": created,
            "updated": updated,
            "errors": errors,
            "total": created + updated + len(errors),
        })

    @action(detail=False, methods=["post"], url_path="bulk-stock")
    def bulk_stock(self, request):
        """Update stock for multiple products and/or variants at once."""
        from django.db import transaction

        products = request.data.get("products", [])
        variants = request.data.get("variants", [])

        if not products and not variants:
            return Response({"detail": "Provide products or variants."}, status=status.HTTP_400_BAD_REQUEST)

        updated = 0
        errors = []

        with transaction.atomic():
            for item in products:
                pid = item.get("id")
                stock = item.get("stock")
                if pid is None or stock is None:
                    errors.append(f"Invalid product entry: {item}")
                    continue
                try:
                    stock = int(stock)
                    if stock < 0:
                        raise ValueError
                except (TypeError, ValueError):
                    errors.append(f"Invalid stock for product {pid}: {item.get('stock')}")
                    continue
                count = Product.objects.filter(id=pid).update(stock=stock)
                if count:
                    updated += 1
                else:
                    errors.append(f"Product {pid} not found")

            for item in variants:
                vid = item.get("id")
                stock = item.get("stock")
                if vid is None or stock is None:
                    errors.append(f"Invalid variant entry: {item}")
                    continue
                try:
                    stock = int(stock)
                    if stock < 0:
                        raise ValueError
                except (TypeError, ValueError):
                    errors.append(f"Invalid stock for variant {vid}: {item.get('stock')}")
                    continue
                count = ProductVariant.objects.filter(id=vid).update(stock=stock)
                if count:
                    updated += 1
                else:
                    errors.append(f"Variant {vid} not found")

        return Response({"updated": updated, "errors": errors})


class ReviewViewSet(viewsets.ModelViewSet):
    """Product reviews. Anyone can read approved reviews.

    Authenticated users can create reviews (one per product) and edit/delete
    their own. Staff can see all reviews regardless of status, filter by
    status via ``?status=``, and approve/reject via dedicated actions.
    """

    serializer_class = ReviewSerializer
    filterset_fields = ("product", "rating", "status")
    ordering_fields = ("created_at", "rating")

    def get_permissions(self):
        if self.action in ("list", "retrieve"):
            return [IsAuthenticatedOrReadOnly()]
        if self.action in ("approve", "reject"):
            return [IsAdminUser()]
        # create/update/delete all require auth; update/delete are further
        # scoped to the author's own reviews in get_queryset.
        return [IsAuthenticated()]

    def get_queryset(self):
        qs = Review.objects.select_related("user", "product")
        user = self.request.user

        # Restrict edits/deletes to the author's own reviews.
        if self.action in ("update", "partial_update", "destroy"):
            return qs.filter(user=user)

        # Non-staff users see only approved reviews (public).
        # Staff can optionally filter by status via ?status=.
        if self.action in ("list", "retrieve"):
            if not (user and user.is_staff):
                qs = qs.filter(status=Review.Status.APPROVED)

        return qs

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        review = self.get_object()
        review.status = Review.Status.APPROVED
        review.save(update_fields=["status"])
        return Response({"detail": "Review approved.", "status": review.status})

    @action(detail=True, methods=["post"])
    def reject(self, request, pk=None):
        review = self.get_object()
        review.status = Review.Status.REJECTED
        review.save(update_fields=["status"])
        return Response({"detail": "Review rejected.", "status": review.status})


class WishlistItemViewSet(viewsets.ModelViewSet):
    """Authenticated user's wishlist. Add/remove/view saved products."""

    serializer_class = WishlistItemSerializer
    permission_classes = [IsAuthenticated]
    http_method_names = ["get", "post", "delete", "head", "options"]

    def get_queryset(self):
        return WishlistItem.objects.filter(user=self.request.user).select_related("product")

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class ProductVariantViewSet(viewsets.ModelViewSet):
    """Variants for a product. Staff can create/update/delete; anyone can read."""
    serializer_class = ProductVariantSerializer
    filterset_fields = ("product", "is_active")

    def get_queryset(self):
        return ProductVariant.objects.select_related("product").all()

    def get_permissions(self):
        if self.action in ("list", "retrieve"):
            return [IsAuthenticatedOrReadOnly()]
        return [IsAdminUser()]


class ProductImageViewSet(viewsets.ModelViewSet):
    """Staff-only: manage additional product images. Anyone can read."""
    serializer_class = ProductImageSerializer
    filterset_fields = ("product",)

    def get_queryset(self):
        return ProductImage.objects.select_related("product").all()

    def get_permissions(self):
        if self.action in ("list", "retrieve"):
            return [IsAuthenticatedOrReadOnly()]
        return [IsAdminUser()]

    def get_parsers(self):
        from rest_framework.parsers import MultiPartParser, JSONParser
        return [MultiPartParser(), JSONParser()]


class WarehouseViewSet(viewsets.ModelViewSet):
    """Staff-only: manage warehouses."""
    queryset = Warehouse.objects.all()
    serializer_class = WarehouseSerializer
    permission_classes = [IsAdminUser]


class WarehouseStockViewSet(viewsets.ModelViewSet):
    """Staff-only: manage stock levels per warehouse."""
    queryset = WarehouseStock.objects.select_related("product", "variant", "warehouse").all()
    serializer_class = WarehouseStockSerializer
    permission_classes = [IsAdminUser]
    filterset_fields = ("warehouse", "product")
